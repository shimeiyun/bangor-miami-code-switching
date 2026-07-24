from __future__ import annotations

import csv
from pathlib import Path

from .utils import (
    completed_years,
    education_group,
    normalize_age,
    normalize_filename,
    normalize_sex,
)


def _standardize(raw: dict, source_row: int) -> dict:
    calculated = completed_years(raw.get("Date of Recording"), raw.get("D.O.B."))
    reported = normalize_age(raw.get("Age") or raw.get("age"))
    matching_age = calculated if calculated is not None else reported
    level = normalize_age(raw.get("Edu Level") or raw.get("education_level"))
    return {
        "source_row": source_row,
        "questionnaire_id": str(
            raw.get("ID") or raw.get("questionnaire_id") or ""
        ).strip(),
        "soundfile": normalize_filename(
            raw.get("Soundfile (*.wav)") or raw.get("soundfile")
        ),
        "sex_normalized": normalize_sex(raw.get("Sex") or raw.get("sex")),
        "matching_age_years": matching_age,
        "matching_age_source": (
            "calculated_from_recording_date_and_dob"
            if calculated is not None
            else ("reported_age" if reported is not None else "")
        ),
        "education_level": level,
        "education_group": education_group(level),
    }


def read_metadata(path: Path) -> list[dict]:
    """Read a portfolio CSV or the original-style questionnaire XLSX."""
    if path.suffix.lower() == ".csv":
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            return [
                _standardize(row, index)
                for index, row in enumerate(csv.DictReader(handle), start=2)
            ]
    if path.suffix.lower() != ".xlsx":
        raise ValueError("Metadata must be .csv or .xlsx")
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError(
            "Reading XLSX requires: pip install 'bangor-miami-codeswitch[xlsx]'"
        ) from exc
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook["Main_Corpus_ Questionnaires"]
    headers = [cell.value for cell in sheet[2]]
    rows = []
    for excel_row, values in enumerate(
        sheet.iter_rows(min_row=3, values_only=True), start=3
    ):
        raw = dict(zip(headers, values))
        if raw.get("ID") or raw.get("Soundfile (*.wav)"):
            rows.append(_standardize(raw, excel_row))
    return rows

